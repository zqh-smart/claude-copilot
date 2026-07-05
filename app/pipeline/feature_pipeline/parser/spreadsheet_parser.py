from __future__ import annotations

from io import BytesIO
from typing import Any

import xlrd
from openpyxl import load_workbook

from app.pipeline.feature_pipeline.parser.helpers import with_parse_metadata
from src.claude_copilot.schemas.document import (
    DocumentMetadata,
    ParsedDocument,
    ParsedPageBlock,
    ParsedSection,
    ParsedTable,
)


class SpreadsheetDocumentParser:
    def parse(self, *, doc_id: str, content: bytes, metadata: DocumentMetadata) -> ParsedDocument:
        extension = (metadata.extension or "").lower()
        if extension == ".xls":
            sheet_payloads = self._load_xls(content)
            parse_backend = "native-xls"
            parse_route = "native_xls"
        else:
            sheet_payloads = self._load_xlsx(content)
            parse_backend = "native-openpyxl"
            parse_route = "native_xlsx"

        layouts = [self._build_sheet_layout(sheet) for sheet in sheet_payloads]
        sections = self._build_sections(doc_id, layouts)
        tables = self._build_tables(doc_id, layouts)
        page_blocks = self._build_page_blocks(doc_id, layouts, tables)
        raw_parts = [section.content for section in sections if section.content.strip()]
        raw_parts.extend(table.raw_markdown or "" for table in tables)
        raw_text = "\n\n".join(part for part in raw_parts if part.strip())

        return ParsedDocument(
            doc_id=doc_id,
            metadata=with_parse_metadata(
                metadata,
                parse_backend=parse_backend,
                parse_route=parse_route,
                page_count=len(sheet_payloads),
                parsed_page_range=(1, len(sheet_payloads)) if sheet_payloads else None,
                parsed_page_count=len(sheet_payloads),
            ),
            raw_text=raw_text,
            sections=sections,
            page_blocks=page_blocks,
            tables=tables,
        )

    def _load_xlsx(self, content: bytes) -> list[dict[str, Any]]:
        workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
        sheets: list[dict[str, Any]] = []
        for sheet_index, worksheet in enumerate(workbook.worksheets, start=1):
            rows = [
                [self._normalize_cell_value(value) for value in row]
                for row in worksheet.iter_rows(values_only=True)
            ]
            sheets.append({"name": worksheet.title, "index": sheet_index, "rows": rows})
        return sheets

    def _load_xls(self, content: bytes) -> list[dict[str, Any]]:
        workbook = xlrd.open_workbook(file_contents=content)
        sheets: list[dict[str, Any]] = []
        for sheet_index, worksheet in enumerate(workbook.sheets(), start=1):
            rows = []
            for row_index in range(worksheet.nrows):
                row = [
                    self._normalize_cell_value(worksheet.cell_value(row_index, col_index))
                    for col_index in range(worksheet.ncols)
                ]
                rows.append(row)
            sheets.append({"name": worksheet.name, "index": sheet_index, "rows": rows})
        return sheets

    def _build_sheet_layout(self, sheet: dict[str, Any]) -> dict[str, Any]:
        elements: list[dict[str, Any]] = []
        current_block: list[list[str]] = []
        current_start_row = 1
        pending_heading: str | None = None
        table_counter = 0

        def flush_block(block: list[list[str]], start_row: int, heading: str | None) -> str | None:
            nonlocal table_counter
            if not block:
                return heading
            if self._looks_like_heading_block(block):
                heading_text = self._collapse_heading_block(block)
                if heading_text:
                    elements.append(
                        {
                            "kind": "heading",
                            "text": heading_text,
                            "row_start": start_row,
                            "row_end": start_row + len(block) - 1,
                        }
                    )
                    return heading_text
                return heading

            table_counter += 1
            elements.append(
                {
                    "kind": "table",
                    "rows": block,
                    "title": heading or str(sheet["name"]),
                    "row_start": start_row,
                    "row_end": start_row + len(block) - 1,
                    "table_order": table_counter,
                }
            )
            return heading

        for row_index, row in enumerate(sheet["rows"], start=1):
            normalized_row = self._trim_trailing_empty_cells(row)
            if any(cell for cell in normalized_row):
                if not current_block:
                    current_start_row = row_index
                current_block.append(normalized_row)
                continue

            pending_heading = flush_block(current_block, current_start_row, pending_heading)
            current_block = []

        pending_heading = flush_block(current_block, current_start_row, pending_heading)

        return {
            "name": sheet["name"],
            "index": sheet["index"],
            "rows": sheet["rows"],
            "elements": elements,
            "last_heading": pending_heading,
        }

    def _build_sections(self, doc_id: str, layouts: list[dict[str, Any]]) -> list[ParsedSection]:
        sections: list[ParsedSection] = []
        for sheet in layouts:
            content_lines = []
            for row in sheet["rows"]:
                cleaned_row = [cell for cell in row if cell]
                if cleaned_row:
                    content_lines.append("\t".join(cleaned_row))

            sections.append(
                ParsedSection(
                    section_id=f"{doc_id}-sheet-{sheet['index']}",
                    title=str(sheet["name"]),
                    content="\n".join(content_lines).strip(),
                    section_type="spreadsheet_sheet",
                    page_start=int(sheet["index"]),
                    page_end=int(sheet["index"]),
                    metadata={"sheet_name": str(sheet["name"]), "sheet_index": int(sheet["index"])},
                )
            )
            for element in sheet["elements"]:
                if element["kind"] != "heading":
                    continue
                sections.append(
                    ParsedSection(
                        section_id=f"{doc_id}-sheet-{sheet['index']}-heading-{len(sections) + 1}",
                        title=element["text"],
                        content=element["text"],
                        section_type="spreadsheet_heading",
                        page_start=int(sheet["index"]),
                        page_end=int(sheet["index"]),
                        metadata={
                            "sheet_name": str(sheet["name"]),
                            "sheet_index": int(sheet["index"]),
                            "source": "spreadsheet_layout",
                            "row_start": int(element["row_start"]),
                            "row_end": int(element["row_end"]),
                        },
                    )
                )
        return sections

    def _build_tables(self, doc_id: str, layouts: list[dict[str, Any]]) -> list[ParsedTable]:
        tables: list[ParsedTable] = []
        for sheet in layouts:
            for element in sheet["elements"]:
                if element["kind"] != "table":
                    continue
                block = element["rows"]
                tables.append(
                    ParsedTable(
                        table_id=f"{doc_id}-sheet-table-{len(tables) + 1}",
                        table_type="spreadsheet_table",
                        title=str(element["title"]),
                        raw_markdown=self._table_to_markdown(block),
                        page=int(sheet["index"]),
                        headers=block[0],
                        rows=block[1:],
                        metadata={
                            "sheet_name": str(sheet["name"]),
                            "sheet_index": int(sheet["index"]),
                            "row_count": len(block),
                            "column_count": max((len(row) for row in block), default=0),
                            "row_start": int(element["row_start"]),
                            "row_end": int(element["row_end"]),
                            "title_source": "preceding_heading" if element["title"] != str(sheet["name"]) else "sheet_name",
                        },
                    )
                )
        return tables

    def _build_page_blocks(
        self,
        doc_id: str,
        layouts: list[dict[str, Any]],
        tables: list[ParsedTable],
    ) -> list[ParsedPageBlock]:
        blocks: list[ParsedPageBlock] = []
        table_key_to_table = {
            (table.page, table.metadata.get("row_start"), table.metadata.get("row_end")): table for table in tables
        }

        for sheet in layouts:
            order = 0
            order += 1
            blocks.append(
                ParsedPageBlock(
                    block_id=f"{doc_id}-sheet-{sheet['index']}-block-{order}",
                    block_type="heading",
                    text=str(sheet["name"]),
                    page=int(sheet["index"]),
                    order=order,
                    metadata={"sheet_name": str(sheet["name"]), "sheet_index": int(sheet["index"])},
                )
            )

            for element in sheet["elements"]:
                order += 1
                if element["kind"] == "heading":
                    blocks.append(
                        ParsedPageBlock(
                            block_id=f"{doc_id}-sheet-{sheet['index']}-block-{order}",
                            block_type="heading",
                            text=str(element["text"]),
                            page=int(sheet["index"]),
                            order=order,
                            metadata={
                                "sheet_name": str(sheet["name"]),
                                "sheet_index": int(sheet["index"]),
                                "row_start": int(element["row_start"]),
                                "row_end": int(element["row_end"]),
                            },
                        )
                    )
                    continue

                table_text = self._table_to_markdown(element["rows"])
                block_id = f"{doc_id}-sheet-{sheet['index']}-block-{order}"
                blocks.append(
                    ParsedPageBlock(
                        block_id=block_id,
                        block_type="table",
                        text=table_text,
                        page=int(sheet["index"]),
                        order=order,
                        metadata={
                            "sheet_name": str(sheet["name"]),
                            "sheet_index": int(sheet["index"]),
                            "row_start": int(element["row_start"]),
                            "row_end": int(element["row_end"]),
                            "table_title": str(element["title"]),
                        },
                    )
                )
                matched_table = table_key_to_table.get((int(sheet["index"]), int(element["row_start"]), int(element["row_end"])))
                if matched_table is not None:
                    matched_table.source_block_id = block_id

        return blocks

    def _split_table_blocks(self, rows: list[list[str]]) -> list[list[list[str]]]:
        blocks: list[list[list[str]]] = []
        current_block: list[list[str]] = []

        for row in rows:
            normalized_row = self._trim_trailing_empty_cells(row)
            if any(cell for cell in normalized_row):
                current_block.append(normalized_row)
                continue

            if current_block:
                blocks.append(current_block)
                current_block = []

        if current_block:
            blocks.append(current_block)

        return blocks

    def _looks_like_heading_block(self, rows: list[list[str]]) -> bool:
        if len(rows) > 2:
            return False

        flattened = [cell.strip() for row in rows for cell in row if cell.strip()]
        if not flattened:
            return False
        if len(flattened) > 3:
            return False

        combined = " ".join(flattened)
        if len(combined) > 120:
            return False
        if sum(1 for value in flattened if self._contains_numeric_signal(value)) >= 2:
            return False
        return True

    def _collapse_heading_block(self, rows: list[list[str]]) -> str:
        values = [cell.strip() for row in rows for cell in row if cell.strip()]
        return " | ".join(values)

    def _contains_numeric_signal(self, value: str) -> bool:
        return any(char.isdigit() for char in value)

    def _normalize_cell_value(self, value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    def _trim_trailing_empty_cells(self, row: list[str]) -> list[str]:
        trimmed = list(row)
        while trimmed and not trimmed[-1]:
            trimmed.pop()
        return trimmed

    @staticmethod
    def _table_to_markdown(rows: list[list[str]]) -> str:
        if not rows:
            return ""
        width = max(len(row) for row in rows)
        normalized = [row + [""] * (width - len(row)) for row in rows]
        header = normalized[0]
        lines = [
            "| " + " | ".join(header) + " |",
            "| " + " | ".join(["---"] * width) + " |",
        ]
        for row in normalized[1:]:
            lines.append("| " + " | ".join(row) + " |")
        return "\n".join(lines)
